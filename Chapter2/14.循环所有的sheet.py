from openpyxl import load_workbook  # 导入三方模块

wb = load_workbook(r"D:\draft\nation.xlsx")  # 打开excel文件

'''for i in wb.sheetnames:  # 循环所有的Sheet名称
    sheet = wb[i]  # 获取当前循环的sheet
    cell = sheet.cell(1, 1)  # 获取当前sheet的某个单元格
    print(cell.value)  # 打印该单元格的内容'''

'''for sheet in wb.worksheets:  # 根据索引进行循环
    cell = sheet.cell(1, 1)  # cell单元格
    print(cell.value)'''  # value内容

for i in wb:  # 直接根据文件进行循环
    cell = i.cell(1, 1)  # 获取cell对应的单元格
    print(cell.value)  # 获取对应的值
