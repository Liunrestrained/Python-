from openpyxl import load_workbook  # 导入三方模块

wb = load_workbook(r"D:\draft\nation.xlsx")  # 打开excel文件
sheet = wb.worksheets[3]

cell = sheet.cell(1, 2)
cell.value = "选手"

wb.save(r"D:\draft\nation.xlsx")  # 将文件保存到……，原路径会覆盖