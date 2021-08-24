from openpyxl import load_workbook  # 导入三方模块

wb = load_workbook(r"D:\draft\nation.xlsx")  # 打开excel文件

print(wb.sheetnames)  # sheet名称

