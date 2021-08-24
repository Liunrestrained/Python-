from openpyxl import load_workbook  # 导入三方模块

wb = load_workbook(r"D:\draft\nation.xlsx")  # 打开excel文件

sheet = wb["China"]  # 选择sheet
cell = sheet.cell(1, 2)  # 选择单元格
print(cell.value)  # 打印单元格内容
