from openpyxl import load_workbook  # 导入三方模块

wb = load_workbook(r"D:\draft\nation.xlsx")  # 打开excel文件

sheet = wb.worksheets[1]

c1 = sheet.cell(1, 1)  # 合并单元格前一个单元格
print(c1)
print(c1.value)  # 能正常打印出来

c2 = sheet.cell(1, 2)  # 合并单元格后一个单元格
print(c2)
print(c2.value)  # None