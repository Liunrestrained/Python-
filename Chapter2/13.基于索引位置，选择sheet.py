from openpyxl import load_workbook  # 导入三方模块

wb = load_workbook(r"D:\draft\nation.xlsx")  # 打开excel文件

sheet = wb.worksheets[0]  # 根据索引选中sheet
cell = sheet.cell(2, 2)  # 选中单元格
print(cell.value)  # 输出单元格的内容
