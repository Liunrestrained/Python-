from openpyxl import load_workbook  # 导入三方模块

wb = load_workbook(r"D:\draft\nation.xlsx")  # 打开excel文件

sheet = wb.worksheets[0]  # 根据索引选中sheet

cell = sheet.cell(1, 1)  # 选中sheet的单元格

print(cell.value)  # 获取内容
print(cell.style)  # 获取样式
print(cell.font)  # 获取字体
print(cell.alignment)  # 获取排列方式
