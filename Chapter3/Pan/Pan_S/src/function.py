'''业务逻辑代码'''
import os  # 路径
import re  #
import json  # 格式
import datetime  # 时间
from openpyxl import load_workbook  # 三方excel模块
from Pan.Pan_S import config
from Pan.Pan_S.utils import method


class Function(object):  # 首先我们将这个类传递给在启动页面传递给了run方法。当有人来访问客户端的时候，就会在run方法中创建一个类的对象。
    # 并且将客户连接conn作为实例变量传递过来，并通过实例化的对象执行类里面的方法。
    def __init__(self, conn):
        self.conn = conn
        self.username = None  # 作为标记，当它为None，就代表连接的用户尚还没有登陆。

    def send_json(self, **kwargs):
        method.send_message(self.conn, json.dumps(kwargs))
        # 这里已经指定了发送对象连接，以及发送方法。后面使用发送转json格式数据时，直接调用这个方法，send_json(内容)，over！

    @property  # 属性方法(绑定方法 + 特殊装饰器 组合创造出来的，在调用方法时可以不加括号)
    def home_path(self):
        return os.path.join(config.file_personal, self.username)  # 路径拼接，当前登录用户的个人文件夹路径

    def recv_file(self, target_file_path):
        method.recv_save_file(self.conn, target_file_path)  # 服务端接收并保存文件

    def send_file_by_seek(self, file_size, file_path, seek=0):  # 提供文件下载
        method.send_file_by_seek(self.conn, file_size, file_path, seek)

    def register(self, username, password):  # *args包含用户名和密码这两个参数
        """注册，并将注册用户名、密码、注册时间，写入用户信息注册表.xlsx文件中"""
        wb = load_workbook(config.file_xlsx)  # 打开xlsx文件
        sheet = wb.worksheets[0]  # 打开第一个Sheet

        """检测用户名是否已经存在于注册表中"""
        exists = False  # 假定用户名不存在
        for row in sheet.iter_rows(2):  # for row in sheet.iter_rows，即：从头开始读取(包含表头)，这里就是从第2行开始向下读单元格
            if username == row[0].value:  # 若用户名和单元格中某个一名字相同 ~ 那就是用户名已存在
                exists = True  # 将exists改为True，也就是用户名存在
                break  # 结束这个for循环，执行下面的告知
        if exists:  # 如果exists为True，也就是当用户名存在时，给客户端回复：用户名已经存在啦！
            # self.conn.sendall(b"……") ==> 粘包
            # method.send_message(self.conn, json.dumps({"status":False, "error":"用户名已存在"}))
            self.send_json(status=False, error="这个用户名好像已经存在了哎！")  # 告知用户注册名已存在
            return  # 方法执行结束，用户可以选择重新注册

        """用户名没毛病，可以注册，具体操作"""
        max_row = sheet.max_row  # 获取注册表已存在数据的行数
        data_list = [username, password, datetime.datetime.now().strftime("%Y-%m_d")]  # 用户名、密码、注册时间(规定年月日格式)
        for i, item in enumerate(data_list, 1):  # 用enumerate标注序号，循环列表中的内容：1.用户名、2.密码、3.注册时间
            cell = sheet.cell(max_row + 1, i)  # 写入位置(已存在数据行数+1，对应列数)
            cell.value = item  # 写入内容
        wb.save(config.file_xlsx)  # 根据本地配置的注册表位置进行保存。

        """创建用户的专属文件目录"""
        user_folder = os.path.join(config.file_personal, username)  # 根据配置文件和用户名拼接出文件路径
        os.makedirs(user_folder)  # 创建这个文件夹路径

        """回复消息，表示注册完工"""
        self.send_json(status=True, error="哇偶！你就是我的第{}位master吗？！真帅！哈哈，我真有福气！".format(max_row))

    def log_in(self, username, password):  # 两个参数：1.用户名、2密码
        """登录，在用户注册表中进行账户校验"""
        wb = load_workbook(config.file_xlsx)  # 打开注册表文件
        sheet = wb.worksheets[0]  # 选定sheet

        online = False  # 在线状态(离线)
        for row in sheet.iter_rows(2):  # 遍历账户的用户名和对应密码
            if username == row[0].value and password == row[1].value:  # 用户名和对应密码正确
                online = True  # 在线状态(在线)
                break  # 结束for循环

        if online:  # 若登录成功
            self.send_json(status=True, error="呜呜呜...Master,你终于回来了，我好想你啊...")  # 告知用户登陆成功
            self.username = username  # 标记修改为用户已经登录

        else:  # 账户名不存在、密码错误、账户名与密码不匹配、违法输入——统称登陆失败
            self.send_json(status=False, error="登陆失败！哼！我一眼就看出来了，你不是我的Master!")  # 告知用户登陆失败

    def ls(self, folder_path=None):  # 一个元素：folder_path,默认为None，用户传文件名或文件夹名时，就为文件(夹)名。
        """查看当前目录下的所有文件
        1.folder_path=None,查看根目录文件；
        2.folder_path不为空，查看用户目录folder_path中的文件。
        """
        if not self.username:  # 当用户还没登陆的时候
            self.send_json(status=False, error="只能对已登录的用户提供服务哦！快登陆吧！😉")
            return  # 不往下走了，结束方法，用户可以选择再次进来

        if not folder_path:  # folder_path=None
            data = "\n".join(os.listdir(self.home_path))  # 返回路径下所有文件的目录，此处使用换行符拼接，将逐行显示。home_path使用了属性，不加()就可以执行。
            self.send_json(status=True, data=data)  # 将文件目录发送给用户
            return  # 不往下走了，结束方法，用户可以选择再次进来

        target_folder = os.path.join(self.home_path, folder_path)  # 路径拼接，用户指定访问的子文件夹

        if not os.path.exists(target_folder):  # 判断路径是否存在
            self.send_json(status=False, error="路径不存在")
            return
        if not os.path.isdir(target_folder):  # 判断文件夹是否存在
            self.send_json(status=False, error="文件夹不存在")
            return

        data = "\n".join(os.listdir(target_folder))  # 如果文件路径和文件夹存在，就把文件目录发过去
        self.send_json(status=True, data=data)

    def up_load(self, file_path):  # 一个元素：文件名
        """上传文件,同名直接覆盖"""
        if not self.username:  # 即self.username=None,用户未登录
            self.send_json(status=False, error="只能对已登录的用户提供服务哦！快登陆吧！😉")
            return

        target_file_path = os.path.join(self.home_path, file_path)  # 要上传的文件路径和文件名
        folder = os.path.dirname(target_file_path)  # 上级文件路径
        if not os.path.exists(folder):  # 上级文件路径如果不存在
            os.makedirs(folder)  # 则创建它

        self.send_json(status=True, error="准备就绪，开始上传")  # 告知用户准备就绪，开始上传文件

        """服务器接收上传的文件"""
        self.recv_file(target_file_path)

    def download(self, file_path, seek=0):  # 续传时：下载的文件和客户端本地文件大小
        """下载文件：要支持断点续传
        seek=None,从头开始下载，本地文件将被覆盖；
        seek=n,从第n个字节开始下载
        """
        if not self.username:  # 离线
            self.send_json(status=False, error="登录成功后才能上传")
            return

        target_file_path = os.path.join(self.home_path, file_path)  # 客户端要下载的文件名和对应的路径
        if not os.path.exists(target_file_path):  # 云盘没有这个文件
            self.send_json(status=False, error="您要下载的文件{}不存在，请检查后重试".format(file_path))
            return

        self.send_json(status=True, data="准备就绪，开始下载")

        seek = int(seek)  # 光标位置，默认为0，续传时，则是本地文件的大小
        total_size = os.stat(target_file_path).st_size  # 网盘本地文件的大小
        method.send_file_by_seek(self.conn, total_size - seek, target_file_path, seek)  # 连接，应该传的大小，文件名字，起始位置

    def Jumping(self):
        """每次客户端来连接，都要执行这个方法。掌管用户的连接与断开：return:False退出循环，断开连接；return:True继续循环，保持连接。"""
        conn = self.conn
        # 1.获取数据包
        cmd = method.receive_message(conn).decode("utf-8")  # (这是一个专门接收信息的方法工具,在method.py文件)接收用户发来的数据信息，并转码。

        if cmd.upper() == "Q":
            print("Master下线了")
            return False  # run做判断，执行break，退出循环，然后断开连接。

        action_rules = {"login": self.log_in,
                        "register": self.register,
                        "ls": self.ls,
                        "upload": self.up_load,
                        "download": self.download,
                        }  # 规定客户的命令对应的指令，作为依据用作系统判定。

        cmd, *args = re.split(r"\s+", cmd)  # 使用正则表达式将用户的命令按照空格进行切割，\s+表示一个或多个空格；若使用字符串的空格，最多只能加2个，有可能导致切割不完整。
        # cmd接收用户的指令，*args则接收一个或者多个命令方法，例如账户和密码。

        order = action_rules[cmd]  # order就是指令后面的方法，加括号就可以执行

        order(*args)  # 方法(数据)，跳到对应位置开始执行它
        return True  # 这个方法运行完了一次，返回信号，循环继续，保持连接
